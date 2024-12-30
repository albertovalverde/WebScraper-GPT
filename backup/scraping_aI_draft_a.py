import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging
from langchain.llms import OpenAI
import os
from urllib.parse import urlparse, urljoin

# Configuración de OpenAI
openai_api_key = os.getenv("OPENAI_API_KEY")
llm = OpenAI(temperature=0, api_key=openai_api_key)

# Estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif !important;
        color: #001978;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Cargar archivo SVG
with open("assets/package_blue.svg", "r") as file:
    icon_svg = file.read()

# Título
st.markdown(f"#### {icon_svg} Buscador Semántico en Sitios Web", unsafe_allow_html=True)

# Subida de archivo y configuración
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")
consulta_input = st.text_area("Describe lo que quieres encontrar", "Necesito encontrar información referente a: denuncia, denuncias, canal de denuncias, canal ético, compliance, Channel, ethics, complaint, canaldenuncias, canaletico, etico, ético, código de conducta, code of conduct, whistleblower channel, Reporting channel, Whistleblowing channel, canal de ética, ética, Complaints Channel, Sistema Interno de Información, Canal del informante, Canal de información, Canal de comunicación interno, General conditions of sale, buen gobierno")

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}

def make_request(url):
    try:
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        response.raise_for_status()
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"Error al acceder a {url}: {e}")
        return None

def dividir_en_fragmentos(texto, max_tokens=1000):
    """Divide el texto en fragmentos pequeños que no excedan el límite de tokens."""
    fragmentos = []
    tokens = texto.split()
    while tokens:
        fragmentos.append(" ".join(tokens[:max_tokens]))
        tokens = tokens[max_tokens:]
    return fragmentos

def buscar_con_ia(texto, consulta, enlaces, url):
    # Dividir el texto en fragmentos más pequeños
    fragmentos = dividir_en_fragmentos(texto, max_tokens=700)
    
    respuestas = []
    enlaces_relevantes = []
    
    # Procesar cada fragmento de texto por separado
    for fragmento in fragmentos:
        prompt = f"""
            Lee el siguiente contenido: "{fragmento}". La consulta es: "{consulta}". 
            También se te proporciona una lista de enlaces de la página web: {', '.join(enlaces[:5])}.
            1. Si el contenido menciona o responde a la consulta, resume la información más relevante en un máximo de 1000 caracteres.
            2. Si es relevante, proporciona enlaces útiles relacionados con la consulta, incluyendo el enlace proporcionado: {url}.
            3. Si no se encuentra información relevante, indica: "No se encontró información relevante relacionada con la consulta".
        """
        
        # Limitar la longitud de la respuesta a 100 tokens
        respuesta = llm(prompt, max_tokens=100)
        respuestas.append(respuesta.strip())
        
        # Filtrar enlaces relevantes (máximo 5 enlaces)
        for enlace in enlaces[:5]:
            if any(palabra in enlace for palabra in consulta.split(", ")):
                enlaces_relevantes.append(enlace)

    # Devolver la respuesta consolidada y los enlaces relevantes
    resultado_final = " ".join(respuestas)
    if "No se encontró información relevante relacionada con la consulta" in resultado_final:
        resultado_icono = "❌ No se encontró información relevante."
    else:
        resultado_icono = "✔️ Se encontró información relevante."

    return resultado_icono, " ".join(respuestas), enlaces_relevantes

def verificar_url(url):
    """Asegura que la URL tenga un esquema válido (http:// o https://)."""
    if not url.startswith("http://") and not url.startswith("https://"):
        return f"https://{url}", f"http://{url}"
    return url, None

def es_url_valida(url):
    if not url or " " in url or not ("." in url):
        return False
    return True

def obtener_enlaces_relevantes(soup, base_url):
    """Función que obtiene los enlaces relevantes de una página web"""
    enlaces = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        # Asegurarse de que los enlaces relativos se resuelvan correctamente
        full_url = urljoin(base_url, href)
        if full_url.startswith('http://') or full_url.startswith('https://'):
            enlaces.append(full_url)
    return enlaces

if st.button("Ejecutar búsqueda") and uploaded_file and selected_column and consulta_input:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        website_column_index = None
        
        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
                break

        if not website_column_index:
            st.error(f"Columna '{selected_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")
        
        # Crear la nueva columna para enlaces relevantes
        enlaces_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=enlaces_col_index, value="Enlaces Relevantes")

        for row in range(2, sheet.max_row + 1):
            url = sheet.cell(row=row, column=website_column_index).value
            if not es_url_valida(url):
                sheet.cell(row=row, column=result_col_index, value="URL inválida o vacía")
                sheet.cell(row=row, column=enlaces_col_index, value="No se encontraron enlaces")
                continue

            # Verificar y corregir la URL
            url_https, url_http = verificar_url(url)
            response = None
            if url_https:
                response = make_request(url_https)
            if not response and url_http:
                response = make_request(url_http)

            if response:
                soup = BeautifulSoup(response.content, "html.parser")
                text = soup.get_text()  # Obtener el texto
                base_url = urlparse(url_https).scheme + "://" + urlparse(url_https).hostname  # Obtener la URL base
                enlaces = obtener_enlaces_relevantes(soup, base_url)  # Obtener todos los enlaces
                resultado_icono, resultado_resumen, enlaces_relevantes = buscar_con_ia(text, consulta_input, enlaces, url_https)
                sheet.cell(row=row, column=result_col_index, value=f"{resultado_icono} {resultado_resumen}")
                sheet.cell(row=row, column=enlaces_col_index, value=", ".join(enlaces_relevantes) if enlaces_relevantes else "No se encontraron enlaces relevantes")
                st.write(f"✔️ Resultado para {url}")
            else:
                sheet.cell(row=row, column=result_col_index, value="Error al acceder")
                sheet.cell(row=row, column=enlaces_col_index, value="Error al acceder")
                st.write(f"❌ Error al acceder a {url}")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
