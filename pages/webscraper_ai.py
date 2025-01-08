import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging
from langchain.llms import OpenAI
import os
from urllib.parse import urlparse, urljoin

# Configuración de OpenAI
openai_api_key = os.getenv("OPENAI_API_KEY")
llm = OpenAI(temperature=0, api_key=openai_api_key)

# Estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif !important;
        color: #001978;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Cargar archivo SVG
with open("assets/package_blue.svg", "r") as file:
    icon_svg = file.read()

# Título
st.markdown(f"#### {icon_svg} Buscador Semántico en Sitios Web", unsafe_allow_html=True)

# Subida de archivo y configuración
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")
consulta_input = st.text_area("Describe lo que quieres encontrar", "Necesito encontrar información referente a: denuncia, denuncias, canal de denuncias, canal ético, compliance, ethics, complaint, canaldenuncias, canaletico, etico, ético, código de conducta, code of conduct, whistleblower channel, Reporting channel, Whistleblowing channel, canal de ética, ética, Complaints Channel, Sistema Interno de Información, Canal del informante, Canal de información, Canal de comunicación interno, General conditions of sale, buen gobierno")

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}

def make_request(url):
    try:
        response = requests.get(url, headers=headers, timeout=15, verify=True)
        response.raise_for_status()
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"Error al acceder a {url}: {e}")
        return None

def verificar_url(url):
    """Asegura que la URL tenga un esquema válido (http:// o https://)."""
    if not url.startswith("http://") and not url.startswith("https://"):
        return f"https://{url}", f"http://{url}"
    return url, None

def es_url_valida(url):
    if not url or " " in url or not ("." in url):
        return False
    return True

def obtener_enlaces_relevantes(soup, base_url, consulta):
    """Función que obtiene los enlaces relevantes de una página web filtrados por la consulta"""
    enlaces_relevantes = []
    palabras_clave = consulta.lower().split(",")  # Extraemos las palabras clave de la consulta

    for a in soup.find_all('a', href=True):
        href = a['href']
        texto = a.get_text().lower()

        # Filtrar los enlaces que contienen palabras clave de la consulta
        if any(palabra.strip() in href.lower() or palabra.strip() in texto for palabra in palabras_clave):
            full_url = urljoin(base_url, href)  # Asegurar que los enlaces relativos se resuelvan
            if full_url.startswith('http://') or full_url.startswith('https://'):
                enlaces_relevantes.append(full_url)

    return enlaces_relevantes

def dividir_en_fragmentos(texto, max_tokens=1000):
    """Divide el texto en fragmentos pequeños que no excedan el límite de tokens."""
    fragmentos = []
    tokens = texto.split()
    while tokens:
        fragmentos.append(" ".join(tokens[:max_tokens]))
        tokens = tokens[max_tokens:]
    return fragmentos

def buscar_con_ia(enlaces, consulta):
    """Función para hacer una búsqueda semántica usando IA sobre los enlaces y la consulta"""
    # Filtramos los enlaces que contienen información relevante a la consulta
    fragmentos = []
    enlaces_relevantes = []

    for enlace in enlaces:
        prompt = f"""
        El siguiente enlace fue encontrado: {enlace}.
        La consulta es: "{consulta}". 
        ¿Este enlace contiene información relacionada con la consulta? Si es así, indícalo.
        """
        respuesta = llm(prompt, max_tokens=100).strip()
        if "sí" in respuesta.lower():
            enlaces_relevantes.append(enlace)

    return enlaces_relevantes

if st.button("Ejecutar búsqueda") and uploaded_file and selected_column and consulta_input:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        website_column_index = None
        
        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
                break

        if not website_column_index:
            st.error(f"Columna '{selected_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")
        
        # Crear la nueva columna para enlaces relevantes
        enlaces_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=enlaces_col_index, value="Enlaces Relevantes")

        # Contador para los logs
        contador = 1

        for row in range(2, sheet.max_row + 1):
            url = sheet.cell(row=row, column=website_column_index).value
            if not es_url_valida(url):
                sheet.cell(row=row, column=result_col_index, value="URL inválida o vacía")
                sheet.cell(row=row, column=enlaces_col_index, value="No se encontraron enlaces relevantes")
                continue

            # Verificar y corregir la URL
            url_https, url_http = verificar_url(url)
            response = None
            if url_https:
                response = make_request(url_https)
            if not response and url_http:
                response = make_request(url_http)

            if response:
                soup = BeautifulSoup(response.content, "html.parser")
                base_url = urlparse(url_https).scheme + "://" + urlparse(url_https).hostname  # Obtener la URL base
                enlaces = obtener_enlaces_relevantes(soup, base_url, consulta_input)  # Obtener solo los enlaces relevantes
                enlaces_relevantes = buscar_con_ia(enlaces, consulta_input)  # Filtrar con IA los enlaces relevantes
                
                # Agregar contador en el log
                if enlaces_relevantes:
                    sheet.cell(row=row, column=result_col_index, value=f"✔️ Enlaces relevantes encontrados ({contador}).")
                    sheet.cell(row=row, column=enlaces_col_index, value=", ".join(enlaces_relevantes))
                    contador += 1
                else:
                    sheet.cell(row=row, column=result_col_index, value=f"❌ No se encontró información relevante ({contador}).")
                    sheet.cell(row=row, column=enlaces_col_index, value="No se encontraron enlaces relevantes")
                    contador += 1
                st.write(f"✔️ Resultado para {url} ({contador - 1})")
            else:
                sheet.cell(row=row, column=result_col_index, value="Error al acceder")
                sheet.cell(row=row, column=enlaces_col_index, value="Error al acceder")
                st.write(f"❌ Error al acceder a {url}")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
