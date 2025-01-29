import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging
from langchain.llms import OpenAI
import os
from urllib.parse import urlparse, urljoin
from duckduckgo_search import DDGS  # Import necesario para generar URL alternativa

# Configuración de OpenAI
openai_api_key = os.getenv("OPENAI_API_KEY")
llm = OpenAI(temperature=0, api_key=openai_api_key)

# Estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif !important;
        color: #001978;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Cargar archivo SVG
with open("assets/package_blue.svg", "r") as file:
    icon_svg = file.read()

# Título
st.markdown(f"#### {icon_svg} Buscador Semántico en Sitios Web", unsafe_allow_html=True)

# Subida de archivo y configuración
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")
empresa_column = st.text_input("Nombre de la columna con razón social", "RAZON_SOCIAL")
consulta_input = st.text_area(
    "Describe lo que quieres encontrar", 
    "Necesito encontrar información referente a: buen gobierno, Canal de comunicación interno, canal de denuncias, canal de ética, Canal de información, Canal de Sistemas de Información, Canal del informante, canal ético, Canal ético, canal-de-denuncias, canal-denuncia, canaldenuncias, Canales Internos de Información, canaletico, code of conduct, código de conducta, Código ético, Codigo_Etico, complaint, Complaints Channel, compliance, denuncia, denuncias, Ethic channel, ethics, ética, etico, ético, General conditions of sale, Reporting channel, Sistema Interno de Información, whistleblower channel, Whistleblowing channel"
)

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}

def make_request(url):
    try:
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        response.raise_for_status()
        logging.info(f"Acceso exitoso a {url}")
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"Error al acceder a {url}: {e}")
        return None


def verificar_url(url, empresa, log_count, row, sheet, url_alternativa_col_index):
    """
    Verifica la URL y genera una alternativa si no es accesible.
    
    Args:
        url (str): URL inicial.
        empresa (str): Razón social de la empresa.
        log_count (int): Contador de la iteración actual.
        row (int): Número de la fila en el Excel.
        sheet (object): Objeto de la hoja de cálculo de openpyxl.
        url_alternativa_col_index (int): Índice de la columna donde se debe guardar la URL alternativa.
        
    Returns:
        str: URL válida o None si no se pudo generar.
    """
    if not url or " " in url or "." not in url:
        url = generar_url_alternativa(empresa)
    
    if not url:
        return None

    if not url.startswith("http://") and not url.startswith("https://"):
        url = f"https://{url}"

    # Intento de acceso a la URL
    response = make_request(url)
    
    if not response:
        # Si falla el acceso, intentamos con DuckDuckGo
        logging.info(f"❌ Error al acceder a la URL original {url}. Intentando obtener una alternativa desde DuckDuckGo.")
        url_alternativa = generar_url_alternativa(empresa)
        
        if url_alternativa:
            logging.info(f"🔄 URL alternativa encontrada: {url_alternativa}. Reintentando acceso.")
            # Mostrar en la interfaz con el contador y el icono al final
            st.write(f"{log_count}. 🔄 URL alternativa generada para {empresa}: {url_alternativa}")
            
            # Actualizar el archivo Excel con la URL alternativa
            sheet.cell(row=row, column=url_alternativa_col_index, value=url_alternativa)
            return url_alternativa
        else:
            logging.error(f"❌ No se pudo generar una URL alternativa para la empresa {empresa}.")
            return None  # Si no se pudo encontrar ninguna alternativa

    return url  # Si el acceso fue exitoso, devolvemos la URL original



def obtener_enlaces_relevantes(soup, base_url, consulta):
    enlaces_relevantes = []
    palabras_clave = consulta.lower().split(",")

    for a in soup.find_all('a', href=True):
        href = a['href']
        texto = a.get_text().lower()

        if any(palabra.strip() in href.lower() or palabra.strip() in texto for palabra in palabras_clave):
            full_url = urljoin(base_url, href)
            if full_url.startswith('http://') or full_url.startswith('https://'):
                enlaces_relevantes.append(full_url)

    return enlaces_relevantes

def buscar_con_ia(enlaces, consulta):
    enlaces_relevantes = []

    for enlace in enlaces:
        prompt = f"""
        El siguiente enlace fue encontrado: {enlace}.
        La consulta es: "{consulta}". 
        ¿Este enlace contiene información relacionada con la consulta? Responde sí o no.
        """
        respuesta = llm(prompt).strip()
        if "sí" in respuesta.lower():
            enlaces_relevantes.append(enlace)

    return enlaces_relevantes

def generar_url_alternativa(empresa):
    """
    Obtiene la primera URL relevante para la empresa utilizando DuckDuckGo.

    Args:
        empresa (str): Nombre o razón social de la empresa.

    Returns:
        str: Primera URL relevante encontrada o None si no hay resultados.
    """
    ddgs = DDGS()
    try:
        resultados = ddgs.text(f"sitio oficial {empresa}", max_results=1)
        if resultados and len(resultados) > 0:
            return resultados[0]['href']  # Acceder al primer elemento de la lista
        return None  # Si no hay resultados
    except Exception as e:
        logging.error(f"Error al buscar URL alternativa: {e}")
        return None


if st.button("Ejecutar búsqueda") and uploaded_file and selected_column and empresa_column and consulta_input:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        website_column_index = None
        empresa_column_index = None

        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
            if cell.value == empresa_column:
                empresa_column_index = cell.column

        if not website_column_index or not empresa_column_index:
            st.error(f"Columna '{selected_column}' o '{empresa_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")

        enlaces_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=enlaces_col_index, value="Enlaces Relevantes")

        # Columna de URL alternativa
        url_alternativa_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=url_alternativa_col_index, value="URL Alternativa")

        log_count = 1

        for row in range(2, sheet.max_row + 1):
            url = sheet.cell(row=row, column=website_column_index).value
            empresa = sheet.cell(row=row, column=empresa_column_index).value

            url_verificada = verificar_url(url, empresa, log_count, row, sheet, url_alternativa_col_index)  # Pasamos los nuevos parámetros

            if not url_verificada:
                # Cuando no se pudo verificar la URL ni encontrar alternativa
                sheet.cell(row=row, column=result_col_index, value="❌ No se pudo verificar o generar URL.")
                sheet.cell(row=row, column=enlaces_col_index, value="No se encontraron enlaces relevantes")
                sheet.cell(row=row, column=url_alternativa_col_index, value="No se generó URL alternativa")
                st.write(f"{log_count}. {empresa}: ❌ No se pudo verificar o generar URL.")
                log_count += 1
                continue

            response = make_request(url_verificada)

            if response:
                soup = BeautifulSoup(response.content, "html.parser")
                base_url = urlparse(url_verificada).scheme + "://" + urlparse(url_verificada).hostname
                enlaces = obtener_enlaces_relevantes(soup, base_url, consulta_input)
                enlaces_relevantes = buscar_con_ia(enlaces, consulta_input)

                if enlaces_relevantes:
                    sheet.cell(row=row, column=result_col_index, value="✔️ Enlaces relevantes encontrados.")
                    sheet.cell(row=row, column=enlaces_col_index, value=", ".join(enlaces_relevantes))
                    st.write(f"{log_count}. {url_verificada}: ✔️ Enlaces relevantes encontrados | {', '.join(enlaces_relevantes)}")
                else:
                    sheet.cell(row=row, column=result_col_index, value="ℹ️ No se encontró información relevante.")
                    sheet.cell(row=row, column=enlaces_col_index, value="ℹ️ No se encontraron enlaces relevantes")
                    st.write(f"{log_count}. {url_verificada}: ℹ️ No se encontró información relevante")
            else:
                # Cuando la URL alternativa tampoco se puede acceder
                st.write(f"{log_count}. {url_verificada}: ❌ Error al acceder incluso con alternativa")
            log_count += 1

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
        logging.exception("Error durante la ejecución de la búsqueda")
