import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging

# Estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif !important;
        color: #001978;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Cargar el archivo SVG desde tu sistema local
with open("assets/package_blue.svg", "r") as file:
    icon_svg = file.read()

# Título con SVG al lado
st.markdown(f"#### {icon_svg} Buscador de Palabras Clave en Sitios Web", unsafe_allow_html=True)

# Entrada de palabras clave
keywords_input = st.text_area("Palabras clave (separadas por comas)", 
    "denuncia, denuncias, canal de denuncias, canal ético, compliance, "
    "Channel, ethics, complaint, canaldenuncias, canaletico, etico, ético, "
    "código de conducta, code of conduct, whistleblower channel, Reporting channel, "
    "Whistleblowing channel, canal de ética, ética, Complaints Channel, "
    "Sistema Interno de Información, Canal del informante, Canal de información, "
    "Canal de comunicación interno, General conditions of sale, buen gobierno")
keywords = [kw.strip().lower() for kw in keywords_input.split(",")]

# Subir archivo Excel
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

# Configuración de columna
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")

# Headers para simular un navegador
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

def make_request(url):
    session = requests.Session()
    session.headers.update(headers)
    try:
        response = session.get(url, timeout=10, verify=False)
        response.raise_for_status()
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"Error al acceder a {url}: {e}")
        return None

# Contador para el log
counter = 0

def verificar_url(url):
    """Verifica si la URL tiene el esquema adecuado. Si no tiene, intenta agregar http:// o https://."""
    if not url.startswith("http://") and not url.startswith("https://"):
        url_https = f"https://{url}"
        url_http = f"http://{url}"
        return url_https, url_http
    return url, None

# Botón de ejecución
if st.button("Ejecutar búsqueda") and uploaded_file and selected_column:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        website_column_index = None
        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
                break

        if not website_column_index:
            st.error(f"Columna '{selected_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")

        for row in range(2, sheet.max_row + 1):
            # Incrementar el contador
            counter += 1
            # Agregar el log con el número de fila procesada y contador
            log_message = f"{counter} - Procesando fila {row - 1} de {sheet.max_row - 1}"
            logging.info(log_message)
            
            url = sheet.cell(row=row, column=website_column_index).value
            if url:
                # Verificar y corregir la URL (http:// o https://)
                url_https, url_http = verificar_url(url)
                response = None
                if url_https:
                    response = make_request(url_https)
                if not response and url_http:
                    response = make_request(url_http)

                if response:
                    soup = BeautifulSoup(response.content, "html.parser")
                    text = soup.get_text().lower()

                    found = False
                    for keyword in keywords:
                        if keyword in text:
                            link = soup.find('a', string=lambda text: text and keyword in text.lower())
                            link_href = link['href'] if link else f"Palabra clave encontrada: '{keyword}' - Link no encontrado"
                            sheet.cell(row=row, column=result_col_index, value=link_href)
                            st.write(f"{counter} ✔️ Palabra clave '{keyword}' encontrada en {url}")
                            found = True
                            break

                    if not found:
                        sheet.cell(row=row, column=result_col_index, value="Palabras clave no encontradas")
                        st.write(f"{counter} ⚠️ No se encontraron palabras clave en {url}")
                else:
                    sheet.cell(row=row, column=result_col_index, value="Error al acceder después de reintentos")
                    st.write(f"{counter} ❌ Error al acceder a {url}")
            else:
                sheet.cell(row=row, column=result_col_index, value="URL vacía")
                st.write(f"{counter} ⚠️ URL vacía en la fila {row}")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
    
